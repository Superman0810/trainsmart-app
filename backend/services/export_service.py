import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

SESSION_COLORS = {
    "run": "FF4CAF50",
    "bike": "FF2196F3",
    "swim": "FF00BCD4",
    "strength": "FFFF9800",
    "mobility": "FF9C27B0",
    "rest": "FFB0BEC5",
}

PHASE_COLORS = {
    "initial": "FFE3F2FD",
    "progression": "FFFFF3E0",
    "taper": "FFF3E5F5",
    "recovery": "FFE8F5E9",
}


def _thin_border():
    s = Side(style="thin", color="FFCCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def generate_excel(sessions: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Training Plan"

    headers = ["Week", "Date", "Day", "Phase", "Type", "Title", "Description", "Duration (min)", "Distance (km)", "Intensity"]
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF263238")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

    ws.row_dimensions[1].height = 20

    for row_idx, s in enumerate(sessions, 2):
        session_date = s["date"] if isinstance(s["date"], date) else date.fromisoformat(str(s["date"]))
        values = [
            s["week_number"],
            session_date.strftime("%Y-%m-%d"),
            session_date.strftime("%A"),
            (s.get("phase") or "").capitalize(),
            (s.get("type") or "").capitalize(),
            s.get("title", ""),
            s.get("description", ""),
            s.get("duration_min"),
            s.get("distance_km"),
            (s.get("intensity") or "").capitalize(),
        ]
        phase = s.get("phase") or "initial"
        session_type = s.get("type") or "rest"
        row_fill = PatternFill("solid", fgColor=PHASE_COLORS.get(phase, "FFFFFFFF"))
        type_fill = PatternFill("solid", fgColor=SESSION_COLORS.get(session_type, "FFFFFFFF"))

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 7))
            if col_idx == 5:
                cell.fill = type_fill
                cell.font = Font(bold=True)
            else:
                cell.fill = row_fill

    col_widths = [8, 12, 12, 12, 12, 30, 60, 16, 15, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_pdf_html(sessions: list[dict], plan_meta: dict) -> str:
    phases = {}
    for s in sessions:
        phase = s.get("phase") or "initial"
        phases.setdefault(phase, []).append(s)

    phase_order = ["initial", "progression", "taper", "recovery"]

    weeks_html = ""
    for phase in phase_order:
        phase_sessions = phases.get(phase, [])
        if not phase_sessions:
            continue

        phase_weeks = {}
        for s in phase_sessions:
            phase_weeks.setdefault(s["week_number"], []).append(s)

        phase_color = {"initial": "#E3F2FD", "progression": "#FFF3E0", "taper": "#F3E5F5", "recovery": "#E8F5E9"}.get(phase, "#FFF")
        weeks_html += f'<div class="phase" style="background:{phase_color}"><h2>{phase.capitalize()} Phase</h2>'

        for week_num in sorted(phase_weeks.keys()):
            week_sessions = sorted(phase_weeks[week_num], key=lambda x: str(x["date"]))
            weeks_html += f'<div class="week"><h3>Week {week_num}</h3><div class="sessions">'
            for s in week_sessions:
                color = SESSION_COLORS.get(s.get("type", "rest"), "FFB0BEC5")[2:]
                session_date = s["date"] if isinstance(s["date"], date) else date.fromisoformat(str(s["date"]))
                meta = []
                if s.get("duration_min"):
                    meta.append(f"{s['duration_min']} min")
                if s.get("distance_km"):
                    meta.append(f"{s['distance_km']} km")
                if s.get("intensity"):
                    meta.append(s["intensity"].capitalize())
                meta_str = " · ".join(meta)
                weeks_html += f"""<div class="session">
                    <div class="session-header" style="background:#{color}">
                        <span class="session-type">{(s.get('type') or '').upper()}</span>
                        <span class="session-date">{session_date.strftime("%a, %b %d")}</span>
                    </div>
                    <div class="session-body">
                        <strong>{s.get('title', '')}</strong>
                        {f'<div class="meta">{meta_str}</div>' if meta_str else ''}
                        <p>{s.get('description', '')}</p>
                    </div>
                </div>"""
            weeks_html += "</div></div>"
        weeks_html += "</div>"

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; color: #333; }}
  h1 {{ color: #263238; border-bottom: 3px solid #263238; padding-bottom: 10px; }}
  .meta-info {{ color: #666; margin-bottom: 30px; }}
  .phase {{ border-radius: 8px; padding: 16px; margin-bottom: 24px; }}
  .phase h2 {{ margin: 0 0 12px 0; color: #263238; font-size: 1.3em; }}
  .week {{ margin-bottom: 16px; }}
  .week h3 {{ margin: 0 0 8px 0; font-size: 1em; color: #555; }}
  .sessions {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 8px; }}
  .session {{ border-radius: 6px; overflow: hidden; border: 1px solid #ddd; background: white; }}
  .session-header {{ padding: 6px 10px; display: flex; justify-content: space-between; align-items: center; }}
  .session-type {{ font-weight: bold; font-size: 0.75em; color: #333; }}
  .session-date {{ font-size: 0.75em; color: #555; }}
  .session-body {{ padding: 8px 10px; font-size: 0.82em; }}
  .session-body strong {{ display: block; margin-bottom: 4px; }}
  .session-body p {{ margin: 4px 0 0 0; color: #666; font-size: 0.95em; }}
  .meta {{ color: #888; font-size: 0.85em; margin-bottom: 4px; }}
  @page {{ size: A4 landscape; margin: 1.5cm; }}
</style>
</head>
<body>
<h1>Training Plan — {plan_meta.get('race_type', '')} ({plan_meta.get('race_date', '')})</h1>
<p class="meta-info">Athlete: {plan_meta.get('name', '')} · Generated: {date.today().strftime("%B %d, %Y")}</p>
{weeks_html}
</body>
</html>"""
